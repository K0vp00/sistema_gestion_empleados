import Empleado from empleado # type: ignore
import Departamento from departamento # type: ignore
import regitroTiempo from registro_tiempo # type: ignore
import proyecto from proyecto # type: ignore
class Informe():
    def __init__(self, id_informe, nombre_informe, id_empleado):
        self.id_informe = id_informe
        self.nombre_informe = nombre_informe
        self.id_empleado = id_empleado


class ExportarDocumento:
    def exportar(self, nombre_informe):
        raise NotImplementedError("Este método debe ser implementado por una subclase")

class ExportarPDF(ExportarDocumento):
    def exportar(self, nombre_informe):
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        # Crear un nuevo archivo PDF
        pdf = canvas.Canvas(nombre_informe, pagesize= letter)
        pdf.setTitle("Informe de Empleados y Proyectos")

        # Títulos
        pdf.setFont("Helvetica-Bold", 16)
        pdf.drawString(100, 750, "Informe de Empleados, Proyectos y Registros de Tiempo")
        
        # Espaciado
        pdf.setFont("Helvetica", 12)
        pdf.drawString(100, 720, "Empleados:")
        
        # Datos de empleados
        employees = [
            {"id": 1, "nombre": "Juan Pérez", "departamento": "IT"},
            {"id": 2, "nombre": "María Gómez", "departamento": "Marketing"},
            {"id": 3, "nombre": "Luis Rodríguez", "departamento": "Recursos Humanos"},
        ]
        
        y_position = 700
        for emp in employees:
            pdf.drawString(100, y_position, f"ID: {emp['id']}, Nombre: {emp['nombre']}, Departamento: {emp['departamento']}")
            y_position -= 20

        # Espacio entre secciones
        y_position -= 10
        pdf.drawString(100, y_position, "Proyectos:")
        
        # Datos de proyectos
        projects = [
            {"id": 101, "nombre": "Proyecto A", "estado": "En progreso"},
            {"id": 102, "nombre": "Proyecto B", "estado": "Finalizado"},
            {"id": 103, "nombre": "Proyecto C", "estado": "Pendiente"},
        ]

        y_position -= 20
        for proj in projects:
            pdf.drawString(100, y_position, f"ID: {proj['id']}, Nombre: {proj['nombre']}, Estado: {proj['estado']}")
            y_position -= 20

        # Espacio entre secciones
        y_position -= 10
        pdf.drawString(100, y_position, "Departamentos:")
        
        # Datos de departamentos
        departments = [
            {"nombre": "IT", "lider": "Juan Pérez"},
            {"nombre": "Marketing", "lider": "María Gómez"},
            {"nombre": "Recursos Humanos", "lider": "Luis Rodríguez"},
        ]
        
        y_position -= 20
        for dept in departments:
            pdf.drawString(100, y_position, f"Nombre: {dept['nombre']}, Líder: {dept['lider']}")
            y_position -= 20

        # Espacio entre secciones
        y_position -= 10
        pdf.drawString(100, y_position, "Registros de Tiempo:")
        
        # Datos de registros de tiempo
        time_records = [
            {"empleado": "Juan Pérez", "proyecto": "Proyecto A", "horas": 8},
            {"empleado": "María Gómez", "proyecto": "Proyecto B", "horas": 6},
            {"empleado": "Luis Rodríguez", "proyecto": "Proyecto C", "horas": 7},
        ]
        
        y_position -= 20
        for record in time_records:
            pdf.drawString(100, y_position, f"Empleado: {record['empleado']}, Proyecto: {record['proyecto']}, Horas: {record['horas']}")
            y_position -= 20

        # Guardar y cerrar el archivo PDF
        pdf.save()
        return f"Exportando '{nombre_informe}' en formato PDF."

class ExportarExcel(ExportarDocumento):
    def exportar(self, nombre_informe):
        import openpyxl
        from openpyxl import Workbook
        # Crear un nuevo archivo de Excel
        wb = Workbook()
        # Datos de empleados
        employees = [
            {"ID": 1, "Nombre": "Juan Pérez", "Departamento": "IT"},
            {"ID": 2, "Nombre": "María Gómez", "Departamento": "Marketing"},
            {"ID": 3, "Nombre": "Luis Rodríguez", "Departamento": "Recursos Humanos"},
        ]

        # Datos de proyectos
        projects = [
            {"ID": 101, "Nombre": "Proyecto A", "Estado": "En progreso"},
            {"ID": 102, "Nombre": "Proyecto B", "Estado": "Finalizado"},
            {"ID": 103, "Nombre": "Proyecto C", "Estado": "Pendiente"},
        ]

        # Datos de departamentos
        departments = [
            {"Nombre": "IT", "Líder": "Juan Pérez"},
            {"Nombre": "Marketing", "Líder": "María Gómez"},
            {"Nombre": "Recursos Humanos", "Líder": "Luis Rodríguez"},
        ]

        # Datos de registros de tiempo
        time_records = [
            {"Empleado": "Juan Pérez", "Proyecto": "Proyecto A", "Horas": 8},
            {"Empleado": "María Gómez", "Proyecto": "Proyecto B", "Horas": 6},
            {"Empleado": "Luis Rodríguez", "Proyecto": "Proyecto C", "Horas": 7},
        ]

        # Crear hojas para cada conjunto de datos
        # Hoja de Empleados
        ws_employees = wb.active
        ws_employees.title = "Empleados"
        ws_employees.append(["ID", "Nombre", "Departamento"])
        for emp in employees:
            ws_employees.append([emp["ID"], emp["Nombre"], emp["Departamento"]])

        # Hoja de Proyectos
        ws_projects = wb.create_sheet(title="Proyectos")
        ws_projects.append(["ID", "Nombre", "Estado"])
        for proj in projects:
            ws_projects.append([proj["ID"], proj["Nombre"], proj["Estado"]])

        # Hoja de Departamentos
        ws_departments = wb.create_sheet(title="Departamentos")
        ws_departments.append(["Nombre", "Líder"])
        for dept in departments:
            ws_departments.append([dept["Nombre"], dept["Líder"]])

        # Hoja de Registros de Tiempo
        ws_time_records = wb.create_sheet(title="Registros de Tiempo")
        ws_time_records.append(["Empleado", "Proyecto", "Horas"])
        for record in time_records:
            ws_time_records.append([record["Empleado"], record["Proyecto"], record["Horas"]])

        # Guardar el archivo de Excel
        wb.save(nombre_informe)    

# Ejemplo de polimorfismo
documentos = [ExportarPDF(), ExportarExcel()]

nombre_informe = "Informe de Ventas"
for documento in documentos:
    print(documento.exportar(nombre_informe))
    
        

